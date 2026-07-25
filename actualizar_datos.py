import os
import json
import re
import html
import unicodedata
from pathlib import Path
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
PLATOS_EDITABLE_PATH = BASE / 'base_platos_editable.xlsx'
LEGACY_EXCEL_PATH = BASE / 'platos_ingredientes.xlsx'
RECIPES_DIR = BASE / 'Recipes'
DATA_JS = BASE / 'js' / 'data.js'

def fix_viewport_in_html(file_path):
    try:
        content = file_path.read_text(encoding='utf-8')
        if '<meta name="viewport"' not in content and '<head>' in content:
            content = content.replace('<head>', '<head>\n        <meta name="viewport" content="width=device-width, initial-scale=1">')
            file_path.write_text(content, encoding='utf-8')
    except Exception:
        pass


def strip_html(text):
    text = re.sub(r'<[^>]+>', ' ', text or '')
    text = html.unescape(text).replace('\xa0', ' ')
    return re.sub(r'\s+', ' ', text).strip()


def parse_recipe_html(file_path):
    content = file_path.read_text(encoding='utf-8', errors='ignore')
    name_match = re.search(r'<h1[^>]*class="name"[^>]*>(.*?)</h1>', content, re.S | re.I)
    cat_match = re.search(r'<p[^>]*class="categories"[^>]*>(.*?)</p>', content, re.S | re.I)
    ingredient_matches = re.findall(r'<p[^>]*class="line"[^>]*itemprop="recipeIngredient"[^>]*>(.*?)</p>', content, re.S | re.I)
    name = strip_html(name_match.group(1)) if name_match else file_path.stem
    raw_category = strip_html(cat_match.group(1)) if cat_match else 'Sin categoría'
    category = normalize_category(raw_category)
    ingredients = []
    for raw in ingredient_matches:
        clean = strip_html(raw)
        clean = re.sub(r'^[-•]\s*', '', clean).strip()
        if clean:
            ingredients.append(clean)
    return {
        'nombre': name,
        'categoria': category,
        'categoria_original': raw_category,
        'oculta_recetario': is_hidden_recipe_category(raw_category),
        'ingredientes_html': ingredients,
        'url': f'Recipes/{file_path.name}'
    }


HIDDEN_RECIPE_CATEGORIES = {'macu', 'findesemana', 'fin de semana', 'finde semana'}


def normalize_category(category):
    parts = [strip_html(p).strip() for p in re.split(r'[,;/]+', category or '')]
    parts = [p for p in parts if p and normalize_text(p) not in HIDDEN_RECIPE_CATEGORIES]
    return parts[0] if len(parts) == 1 else ', '.join(parts) if parts else 'Sin categoría'


def is_hidden_recipe_category(category):
    parts = [strip_html(p).strip() for p in re.split(r'[,;/]+', category or '')]
    visible = [p for p in parts if p and normalize_text(p) not in HIDDEN_RECIPE_CATEGORIES]
    return not visible

def normalize_text(value):
    text = str(value or '').strip().lower()
    text = ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
    replacements = {
        'air fryer': 'airfryer',
        '&': ' y ',
        '/': ' ',
        ',': ' ',
        '(': ' ',
        ')': ' ',
        '¿': ' ',
        '?': ' ',
        '…': ' ',
        '.': ' ',
        '-': ' '
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def token_set(value):
    stopwords = {'de', 'del', 'la', 'el', 'los', 'las', 'con', 'y', 'en', 'al', 'a', 'o', 'estilo'}
    return {t for t in normalize_text(value).split() if t and t not in stopwords}


def pick_recipe_url(plato_nombre, recetas):
    nombre_norm = normalize_text(plato_nombre)
    nombre_tokens = token_set(plato_nombre)

    exact = next((r['url'] for r in recetas if normalize_text(r['nombre']) == nombre_norm), None)
    if exact:
        return exact

    candidates = []
    for receta in recetas:
        receta_norm = normalize_text(receta['nombre'])
        receta_tokens = token_set(receta['nombre'])
        overlap = len(nombre_tokens & receta_tokens)
        if not overlap:
            continue
        dish_coverage = overlap / max(len(nombre_tokens), 1)
        recipe_coverage = overlap / max(len(receta_tokens), 1)
        if len(nombre_tokens) == 1:
            if not (nombre_norm == receta_norm or receta_norm.startswith(nombre_norm + ' ')):
                continue
        elif dish_coverage < 0.75:
            continue
        subset_bonus = 0
        if nombre_norm and nombre_norm in receta_norm:
            subset_bonus = 3
        elif receta_norm and receta_norm in nombre_norm:
            subset_bonus = 2
        score = (dish_coverage * 4) + (recipe_coverage * 2) + overlap + subset_bonus
        candidates.append((score, -overlap, len(receta_tokens), receta['url']))

    if not candidates:
        return None

    candidates.sort(key=lambda x: (-x[0], x[2]))
    best_score = candidates[0][0]
    if best_score < 3:
        return None
    return candidates[0][3]


def main():
    print("Iniciando actualización de datos para la Web App...")
    
    # Load Excel
    excel_path = PLATOS_EDITABLE_PATH if PLATOS_EDITABLE_PATH.exists() else LEGACY_EXCEL_PATH
    platos = []
    
    if excel_path.exists():
        print(f"Leyendo base de platos desde: {excel_path.name}")
        wb = load_workbook(excel_path, data_only=True)
        ws = wb['Platos'] if 'Platos' in wb.sheetnames else wb.active
        encabezados = [str(c.value).strip().lower() if c.value is not None else '' for c in ws[1]]
        idx_plato = encabezados.index('plato') if 'plato' in encabezados else 0
        idx_categoria = encabezados.index('categoria') if 'categoria' in encabezados else 1
        
        idx_ingredientes = [i for i, h in enumerate(encabezados) if h.startswith('ingrediente_')]
        if not idx_ingredientes:
            idx_ingredientes = list(range(2, len(encabezados)))
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or idx_plato >= len(row) or not row[idx_plato]:
                continue
            nombre = str(row[idx_plato]).strip()
            categoria = str(row[idx_categoria]).strip() if idx_categoria < len(row) and row[idx_categoria] else 'Sin categoría'
            ingredientes = [str(row[i]).strip() for i in idx_ingredientes if i < len(row) and row[i] not in (None, '')]
            platos.append({
                'plato': nombre,
                'categoria': normalize_category(categoria),
                'ingredientes': ingredientes,
                'oculta_recetario': is_hidden_recipe_category(categoria),
                'en_excel': True
            })
    else:
        print("Advertencia: No se encontró ningún archivo Excel de platos.")

    # Read Recipes
    recetas = []
    if RECIPES_DIR.exists():
        for f in os.listdir(RECIPES_DIR):
            if f.endswith('.html'):
                file_path = RECIPES_DIR / f
                fix_viewport_in_html(file_path)
                recetas.append(parse_recipe_html(file_path))
        recetas.sort(key=lambda x: x['nombre'].lower())
        print(f"Se han encontrado {len(recetas)} recetas en la carpeta Recipes.")

    # Match / merge Excel plates with Paprika recipes
    platos_by_name = {normalize_text(p['plato']): p for p in platos}

    for p in platos:
        matched_url = pick_recipe_url(p['plato'], recetas)
        if matched_url:
            receta = next((r for r in recetas if r['url'] == matched_url), None)
            p['url_receta'] = matched_url
            if receta and not p.get('oculta_recetario'):
                p['oculta_recetario'] = False

    # Add recipes that only exist in Paprika/HTML
    used_recipe_urls = {p.get('url_receta') for p in platos if p.get('url_receta')}
    for receta in recetas:
        key = normalize_text(receta['nombre'])
        if key in platos_by_name or receta['url'] in used_recipe_urls:
            continue
        platos.append({
            'plato': receta['nombre'],
            'categoria': normalize_category(receta.get('categoria') or 'Sin categoría'),
            'ingredientes': receta.get('ingredientes_html', []),
            'url_receta': receta['url'],
            'oculta_recetario': receta.get('oculta_recetario', False),
            'en_excel': False
        })

    platos.sort(key=lambda x: (x.get('categoria', 'Sin categoría').lower(), x.get('plato', '').lower()))
            
    # Write to data.js
    DATA_JS.parent.mkdir(exist_ok=True)
    js_content = f"const platosData = {json.dumps(platos, ensure_ascii=False, indent=2)};\n"
    js_content += f"const recetasData = {json.dumps(recetas, ensure_ascii=False, indent=2)};\n"
    
    DATA_JS.write_text(js_content, encoding='utf-8')
    print(f"✅ Se han exportado {len(platos)} elementos combinados (Excel + Paprika).")
    print(f"✅ Datos guardados correctamente en {DATA_JS}")
    print("--------------------------------------------------")
    print("Sube los cambios a GitHub para actualizar la web.")

if __name__ == '__main__':
    main()
