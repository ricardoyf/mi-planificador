import json
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from pathlib import Path
from openpyxl import load_workbook, Workbook
import subprocess
import html
import os
import datetime
import unicodedata

BASE = Path(__file__).resolve().parent
LEGACY_EXCEL_PATH = BASE / 'platos_ingredientes.xlsx'
PLATOS_EDITABLE_PATH = BASE / 'base_platos_editable.xlsx'
PLAN_JSON = BASE / 'plan_semanal_base.json'
PLAN_PAR_JSON = BASE / 'plan_semanal_par.json'
PLAN_IMPAR_JSON = BASE / 'plan_semanal_impar.json'
PLAN_PLANIFICACION_JSON = BASE / 'plan_semanal_planificacion.json'
CONFIG_JSON = BASE / 'planificador_comidas_config.json'

DIAS = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']
TIPOS = ['comida', 'cena']
SLOTS = ['primero', 'segundo']
COLUMNAS_EXCEL = ['plato', 'categoria', 'ingrediente_1', 'ingrediente_2', 'ingrediente_3', 'ingrediente_4', 'ingrediente_5', 'ingrediente_6', 'ingrediente_7', 'ingrediente_8', 'ingrediente_9', 'ingrediente_10', 'ingrediente_11', 'ingrediente_12']
COMPRA_CANONICAL_MAP = {
    'berenjenas': 'Berenjena',
    'patatas': 'Patata',
    'puerros': 'Puerro',
    'tomates': 'Tomate',
    'zanahorias': 'Zanahoria',
}


def normalizar_texto_compra(texto):
    limpio = str(texto or '').strip()
    key = ''.join(
        c for c in unicodedata.normalize('NFD', limpio.lower())
        if unicodedata.category(c) != 'Mn'
    )
    key = ' '.join(key.split())
    return COMPRA_CANONICAL_MAP.get(key, limpio)


class PlanificadorComidas:
    def __init__(self, root):
        self.root = root
        self.root.title('Planificador semanal de comidas')
        self.root.configure(bg='white')
        self.root.update_idletasks()
        screen_w = self.root.winfo_screenwidth()
        screen_h = self.root.winfo_screenheight()
        win_w = min(1880, max(1280, screen_w - 40))
        win_h = min(920, max(760, screen_h - 120))
        pos_x = max(0, (screen_w - win_w) // 2)
        pos_y = max(20, (screen_h - win_h) // 2)
        self.root.geometry(f'{win_w}x{win_h}+{pos_x}+{pos_y}')
        self.root.minsize(1280, 720)
        self.root.deiconify()
        self.root.lift()
        self.root.after(200, lambda: self.root.focus_force())

        self.is_fullscreen = False
        self.root.bind('<F11>', self.toggle_fullscreen)
        self.root.bind('<Escape>', self.end_fullscreen)

        self.config = self.cargar_config()
        self.asegurar_base_platos_editable()
        self.excel_path = self.obtener_ruta_excel_activa()

        self.platos = self.cargar_platos_excel()
        self.categorias = self.obtener_categorias()
        self.plan_path = PLAN_JSON
        self.plan = self.cargar_plan_json()
        self.celdas = {}
        self.plato_seleccionado = None
        self.ultimo_slot = None
        self.categorias_expandida = set(self.categorias)
        self.filtro_texto = tk.StringVar(value='')

        self.setup_ui()
        self.refrescar_arbol_platos()
        self.refrescar_celdas()
        self.refrescar_resumen()

    def toggle_fullscreen(self, event=None):
        self.is_fullscreen = not self.is_fullscreen
        self.root.attributes("-fullscreen", self.is_fullscreen)

    def end_fullscreen(self, event=None):
        if self.is_fullscreen:
            self.is_fullscreen = False
            self.root.attributes("-fullscreen", False)

    def cargar_config(self):
        if CONFIG_JSON.exists():
            try:
                with open(CONFIG_JSON, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    return data
            except Exception:
                pass
        return {}

    def guardar_config(self):
        with open(CONFIG_JSON, 'w', encoding='utf-8') as f:
            json.dump(self.config, f, ensure_ascii=False, indent=2)

    def obtener_ruta_excel_activa(self):
        ruta = self.config.get('excel_path')
        if ruta and Path(ruta).exists():
            return Path(ruta)
        return PLATOS_EDITABLE_PATH if PLATOS_EDITABLE_PATH.exists() else LEGACY_EXCEL_PATH

    def asegurar_base_platos_editable(self):
        if PLATOS_EDITABLE_PATH.exists():
            return

        wb = Workbook()
        ws = wb.active
        ws.title = 'Platos'
        ws.append(COLUMNAS_EXCEL)

        if LEGACY_EXCEL_PATH.exists():
            try:
                wb_old = load_workbook(LEGACY_EXCEL_PATH)
                if 'Platos' in wb_old.sheetnames:
                    ws_old = wb_old['Platos']
                    for row in ws_old.iter_rows(min_row=2, values_only=True):
                        if not row or not row[0]:
                            continue
                        fila = [row[i] if i < len(row) else '' for i in range(len(COLUMNAS_EXCEL))]
                        ws.append(fila)
            except Exception:
                pass

        instrucciones = wb.create_sheet('Instrucciones')
        instrucciones['A1'] = 'Base editable de platos para el planificador.'
        instrucciones['A2'] = 'Edita solo la hoja Platos.'
        instrucciones['A3'] = 'Columnas obligatorias: plato, categoria.'
        instrucciones['A4'] = 'Puedes añadir o quitar filas. Luego pulsa "Recargar base" en la GUI.'
        instrucciones['A5'] = 'Puedes cambiar a otro Excel desde el botón "Elegir otro Excel".'

        wb.save(PLATOS_EDITABLE_PATH)

    def normalizar_plan_legacy(self, data):
        plan = {dia: {tipo: {slot: '' for slot in SLOTS} for tipo in TIPOS} for dia in DIAS}
        for dia in DIAS:
            if dia not in data:
                continue
            valor_dia = data[dia]
            if isinstance(valor_dia, dict):
                for tipo in TIPOS:
                    valor_tipo = valor_dia.get(tipo, '')
                    if isinstance(valor_tipo, dict):
                        for slot in SLOTS:
                            plan[dia][tipo][slot] = valor_tipo.get(slot, '')
                    else:
                        plan[dia][tipo]['primero'] = valor_tipo or ''
                        plan[dia][tipo]['segundo'] = ''
        return plan

    def cargar_plan_json(self, ruta=None):
        ruta_carga = ruta if ruta else self.plan_path
        if ruta_carga.exists():
            with open(ruta_carga, 'r', encoding='utf-8') as f:
                data = json.load(f)
        else:
            data = {}
        return self.normalizar_plan_legacy(data)

    def guardar_plan_json(self, ruta=None):
        ruta_guardado = ruta if ruta else self.plan_path
        with open(ruta_guardado, 'w', encoding='utf-8') as f:
            json.dump(self.plan, f, ensure_ascii=False, indent=2)

    def cargar_platos_excel(self):
        platos = []
        if not self.excel_path.exists():
            return platos

        wb = load_workbook(self.excel_path, data_only=True)
        if 'Platos' in wb.sheetnames:
            ws = wb['Platos']
        else:
            ws = wb.active
        encabezados = [str(c.value).strip().lower() if c.value is not None else '' for c in ws[1]]
        idx_plato = encabezados.index('plato') if 'plato' in encabezados else 0
        idx_categoria = encabezados.index('categoria') if 'categoria' in encabezados else 1
        idx_ingredientes = [i for i, h in enumerate(encabezados) if h.startswith('ingrediente_')]
        if not idx_ingredientes:
            idx_ingredientes = list(range(2, len(encabezados)))

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or idx_plato >= len(row) or not row[idx_plato]:
                continue
            nombre = str(row[idx_plato]).strip()
            categoria = str(row[idx_categoria]).strip() if idx_categoria < len(row) and row[idx_categoria] else 'Sin categoría'
            ingredientes = [str(row[i]).strip() for i in idx_ingredientes if i < len(row) and row[i] not in (None, '')]
            platos.append({
                'plato': nombre,
                'categoria': categoria,
                'ingredientes': ingredientes,
            })
        return platos

    def obtener_categorias(self):
        return sorted({p['categoria'] for p in self.platos})

    def setup_ui(self):
        status_bar = tk.Frame(self.root, bg='#f0f0f0', bd=1, relief='sunken')
        status_bar.pack(side='bottom', fill='x')
        self.lbl_excel = tk.Label(status_bar, text=f'Base actual: {self.excel_path} | F11: Pantalla completa | Esc: Salir pantalla completa', bg='#f0f0f0', fg='#333', anchor='w')
        self.lbl_excel.pack(side='left', padx=8, pady=2)

        top = tk.Frame(self.root, bg='white', pady=6)
        top.pack(fill='x', padx=4)

        tk.Label(top, text='Planificador', bg='white', font=('Arial', 12, 'bold')).pack(side='left', padx=(4, 10))

        tk.Button(top, text='Abrir Excel', command=self.abrir_base_platos).pack(side='left', padx=2)
        tk.Button(top, text='Recargar', command=self.recargar_excel).pack(side='left', padx=2)
        tk.Button(top, text='Copiar lista', command=self.copiar_lista_compra).pack(side='left', padx=2)
        tk.Button(top, text='Exp. HTML', command=self.exportar_html).pack(side='left', padx=2)
        tk.Button(top, text='Exp. PDF', command=self.exportar_pdf).pack(side='left', padx=2)
        tk.Label(top, text='|', bg='white', fg='#ccc').pack(side='left', padx=4)
        tk.Button(top, text='Cargar PAR', command=lambda: self.cargar_plan_directo(PLAN_PAR_JSON)).pack(side='left', padx=2)
        tk.Button(top, text='Cargar IMPAR', command=lambda: self.cargar_plan_directo(PLAN_IMPAR_JSON)).pack(side='left', padx=2)
        tk.Button(top, text='Guardar actual', command=self.guardar_plan).pack(side='left', padx=2)
        tk.Button(top, text='Guardar PAR', command=lambda: self.guardar_plan_como(PLAN_PAR_JSON)).pack(side='left', padx=2)
        tk.Button(top, text='Guardar IMPAR', command=lambda: self.guardar_plan_como(PLAN_IMPAR_JSON)).pack(side='left', padx=2)
        tk.Label(top, text='|', bg='white', fg='#ccc').pack(side='left', padx=4)
        tk.Button(top, text='Expandir', command=self.expandir_todo).pack(side='left', padx=2)
        tk.Button(top, text='Contraer', command=self.contraer_todo).pack(side='left', padx=2)
        tk.Button(top, text='Quitar sel.', command=self.quitar_plato_seleccionado).pack(side='left', padx=2)
        tk.Button(top, text='Comida→Cena', command=self.duplicar_comida_a_cena).pack(side='left', padx=2)
        tk.Button(top, text='Limpiar sem.', command=self.limpiar_semana).pack(side='left', padx=2)

        main = tk.Frame(self.root, bg='white')
        main.pack(fill='both', expand=True, padx=8, pady=6)

        left = tk.Frame(main, bg='white', bd=1, relief='solid', padx=8, pady=8)
        left.pack(side='left', fill='both')

        tk.Label(left, text='Base de platos', bg='white', font=('Arial', 12, 'bold')).pack(anchor='w')
        buscador = tk.Frame(left, bg='white')
        buscador.pack(fill='x', pady=(6, 4))
        tk.Label(buscador, text='Buscar', bg='white').pack(anchor='w')
        ent = tk.Entry(buscador, textvariable=self.filtro_texto)
        ent.pack(fill='x', pady=(1, 0))
        ent.bind('<KeyRelease>', lambda e: self.refrescar_arbol_platos())

        tree_wrap = tk.Frame(left, bg='white')
        tree_wrap.pack(fill='both', expand=True, pady=(2, 6))
        self.tree_platos = ttk.Treeview(tree_wrap, show='tree', height=30)
        scroll = ttk.Scrollbar(tree_wrap, orient='vertical', command=self.tree_platos.yview)
        self.tree_platos.configure(yscrollcommand=scroll.set)
        self.tree_platos.pack(side='left', fill='both', expand=True)
        scroll.pack(side='right', fill='y')
        self.tree_platos.bind('<<TreeviewSelect>>', lambda e: self.seleccionar_plato_desde_arbol())
        self.tree_platos.bind('<Double-1>', lambda e: self.toggle_categoria_seleccionada())

        self.lbl_plato = tk.Label(left, text='Plato seleccionado: ninguno', bg='white', fg='#0b57d0', wraplength=280, justify='left')
        self.lbl_plato.pack(anchor='w', pady=(4, 6))



        center = tk.Frame(main, bg='white')
        center.pack(side='left', fill='both', expand=True, padx=(12, 0))

        tabla = tk.Frame(center, bg='white')
        tabla.pack(fill='x')

        tk.Label(tabla, text='', width=10, bg='white').grid(row=0, column=0, padx=2, pady=2)
        col = 1
        for dia in DIAS:
            tk.Label(tabla, text=dia.capitalize(), bg='white', font=('Arial', 10, 'bold'), width=14).grid(row=0, column=col, columnspan=2, padx=2, pady=2)
            tk.Label(tabla, text='1º', bg='white', fg='#555', font=('Arial', 9, 'bold'), width=7).grid(row=1, column=col, padx=1, pady=(0, 2))
            tk.Label(tabla, text='2º', bg='white', fg='#555', font=('Arial', 9, 'bold'), width=7).grid(row=1, column=col + 1, padx=1, pady=(0, 2))
            col += 2

        base_row = 2
        for idx_tipo, tipo in enumerate(TIPOS):
            tk.Label(tabla, text=tipo.capitalize(), bg='white', font=('Arial', 10, 'bold')).grid(row=base_row + idx_tipo, column=0, padx=2, pady=3, sticky='w')
            col = 1
            for dia in DIAS:
                for slot in SLOTS:
                    frame = tk.Frame(tabla, bg='white', bd=1, relief='solid', width=104, height=92)
                    frame.grid(row=base_row + idx_tipo, column=col, padx=1, pady=3, sticky='nsew')
                    frame.grid_propagate(False)

                    lbl_slot = tk.Label(frame, text='1º plato' if slot == 'primero' else '2º plato', bg='white', fg='#666', font=('Arial', 8, 'bold'))
                    lbl_slot.place(relx=0.5, rely=0.15, anchor='center')

                    lbl = tk.Label(frame, text='', bg='white', wraplength=86, justify='center')
                    lbl.place(relx=0.5, rely=0.48, anchor='center')

                    btns = tk.Frame(frame, bg='white')
                    btns.place(relx=0.5, rely=0.83, anchor='center')
                    tk.Button(btns, text='Poner', width=6, command=lambda d=dia, t=tipo, s=slot: self.asignar_a_celda(d, t, s)).pack(side='left', padx=1)
                    tk.Button(btns, text='Quitar', width=6, command=lambda d=dia, t=tipo, s=slot: self.quitar_de_celda(d, t, s)).pack(side='left', padx=1)

                    frame.bind('<Button-1>', lambda e, d=dia, t=tipo, s=slot: self.asignar_a_celda(d, t, s))
                    lbl.bind('<Button-1>', lambda e, d=dia, t=tipo, s=slot: self.asignar_a_celda(d, t, s))
                    lbl_slot.bind('<Button-1>', lambda e, d=dia, t=tipo, s=slot: self.asignar_a_celda(d, t, s))
                    frame.bind('<Button-3>', lambda e, d=dia, t=tipo, s=slot: self.quitar_de_celda(d, t, s))
                    lbl.bind('<Button-3>', lambda e, d=dia, t=tipo, s=slot: self.quitar_de_celda(d, t, s))
                    lbl_slot.bind('<Button-3>', lambda e, d=dia, t=tipo, s=slot: self.quitar_de_celda(d, t, s))
                    self.celdas[(dia, tipo, slot)] = lbl
                    col += 1

        bottom = tk.Frame(center, bg='white')
        bottom.pack(fill='both', expand=True, pady=(18, 0))

        left2 = tk.Frame(bottom, bg='white')
        left2.pack(side='left', fill='both', expand=True)
        right2 = tk.Frame(bottom, bg='white')
        right2.pack(side='left', fill='both', expand=True, padx=(15, 0))

        tk.Label(left2, text='Resumen semanal', bg='white', font=('Arial', 12, 'bold')).pack(anchor='w')
        self.txt_resumen = tk.Text(left2, height=16, wrap='word')
        self.txt_resumen.pack(fill='both', expand=True)

        tk.Label(right2, text='Lista de la compra estimada', bg='white', font=('Arial', 12, 'bold')).pack(anchor='w')
        self.txt_compra = tk.Text(right2, height=16, wrap='word')
        self.txt_compra.pack(fill='both', expand=True)

    def refrescar_arbol_platos(self):
        self.tree_platos.delete(*self.tree_platos.get_children())
        platos_por_categoria = {categoria: [] for categoria in self.categorias}
        filtro = self.filtro_texto.get().strip().lower()
        for plato in sorted(self.platos, key=lambda x: (x['categoria'], x['plato'])):
            texto_busqueda = ' '.join([
                plato['plato'],
                plato['categoria'],
                ' '.join(plato.get('ingredientes', []))
            ]).lower()
            if filtro and filtro not in texto_busqueda:
                continue
            platos_por_categoria.setdefault(plato['categoria'], []).append(plato)

        for categoria in sorted(platos_por_categoria):
            if not platos_por_categoria[categoria]:
                continue
            abierto = True if filtro else categoria in self.categorias_expandida
            cat_id = self.tree_platos.insert('', 'end', text=categoria, open=abierto, values=('categoria',))
            for plato in platos_por_categoria[categoria]:
                texto = plato['plato']
                self.tree_platos.insert(cat_id, 'end', text=texto, values=('plato', plato['plato']))

    def seleccionar_plato_desde_arbol(self):
        sel = self.tree_platos.selection()
        if not sel:
            return
        item = sel[0]
        values = self.tree_platos.item(item, 'values')
        if not values:
            return
        if values[0] == 'categoria':
            categoria = self.tree_platos.item(item, 'text')
            abierto = self.tree_platos.item(item, 'open')
            self.tree_platos.item(item, open=not abierto)
            if abierto:
                self.categorias_expandida.discard(categoria)
            else:
                self.categorias_expandida.add(categoria)
            return
        self.plato_seleccionado = values[1]
        self.lbl_plato.config(text=f'Plato seleccionado: {self.plato_seleccionado}')

    def toggle_categoria_seleccionada(self):
        sel = self.tree_platos.selection()
        if not sel:
            return
        item = sel[0]
        values = self.tree_platos.item(item, 'values')
        if values and values[0] == 'categoria':
            categoria = self.tree_platos.item(item, 'text')
            abierto = self.tree_platos.item(item, 'open')
            self.tree_platos.item(item, open=not abierto)
            if abierto:
                self.categorias_expandida.discard(categoria)
            else:
                self.categorias_expandida.add(categoria)

    def expandir_todo(self):
        for item in self.tree_platos.get_children():
            categoria = self.tree_platos.item(item, 'text')
            self.tree_platos.item(item, open=True)
            self.categorias_expandida.add(categoria)

    def contraer_todo(self):
        for item in self.tree_platos.get_children():
            categoria = self.tree_platos.item(item, 'text')
            self.tree_platos.item(item, open=False)
            self.categorias_expandida.discard(categoria)

    def buscar_plato(self, nombre):
        for p in self.platos:
            if p['plato'] == nombre:
                return p
        return None

    def refrescar_celdas(self):
        for dia in DIAS:
            for tipo in TIPOS:
                for slot in SLOTS:
                    valor = self.plan[dia][tipo][slot]
                    self.celdas[(dia, tipo, slot)].config(text=valor if valor else '-')

    def asignar_a_celda(self, dia, tipo, slot):
        self.ultimo_slot = (dia, tipo, slot)
        if not self.plato_seleccionado:
            messagebox.showinfo('Selecciona un plato', 'Primero elige un plato de la columna izquierda.')
            return
        self.plan[dia][tipo][slot] = self.plato_seleccionado
        self.refrescar_celdas()
        self.refrescar_resumen()

    def quitar_de_celda(self, dia, tipo, slot):
        self.ultimo_slot = (dia, tipo, slot)
        self.plan[dia][tipo][slot] = ''
        self.refrescar_celdas()
        self.refrescar_resumen()

    def quitar_plato_seleccionado(self):
        self.plato_seleccionado = None
        self.lbl_plato.config(text='Plato seleccionado: ninguno')
        self.tree_platos.selection_remove(self.tree_platos.selection())

    def limpiar_semana(self):
        if not messagebox.askyesno('Confirmar', '¿Quieres limpiar toda la semana?'):
            return
        for dia in DIAS:
            for tipo in TIPOS:
                for slot in SLOTS:
                    self.plan[dia][tipo][slot] = ''
        self.refrescar_celdas()
        self.refrescar_resumen()

    def duplicar_comida_a_cena(self):
        if not messagebox.askyesno('Confirmar', '¿Quieres copiar los platos de comida a cena en toda la semana?'):
            return
        for dia in DIAS:
            self.plan[dia]['cena']['primero'] = self.plan[dia]['comida']['primero']
            self.plan[dia]['cena']['segundo'] = self.plan[dia]['comida']['segundo']
        self.refrescar_celdas()
        self.refrescar_resumen()

    def guardar_plan(self):
        fecha_actual = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        carpeta_json = BASE / 'json'
        carpeta_json.mkdir(exist_ok=True)
        ruta_historico = carpeta_json / f'plan_semanal_{fecha_actual}.json'
        self.guardar_plan_json(ruta_historico)
        
        messagebox.showinfo('Guardado', f'Histórico guardado en:\n{ruta_historico}')

    def guardar_plan_como(self, ruta):
        self.plan_path = ruta
        self.guardar_plan_json()
        messagebox.showinfo('Guardado', f'Plan guardado en:\n{ruta}')

    def cargar_plan_directo(self, ruta):
        if not ruta.exists():
            messagebox.showwarning('Atención', f'El plan no existe:\n{ruta.name}\n\nGuárdalo primero antes de intentar cargarlo.')
            return
        self.plan_path = ruta
        self.plan = self.cargar_plan_json(ruta)
        self.refrescar_celdas()
        self.refrescar_resumen()
        messagebox.showinfo('Cargado', f'Se ha cargado el plan:\n{ruta.name}')

    def abrir_base_platos(self):
        ruta = self.excel_path
        try:
            if os.name == 'nt':
                os.startfile(str(ruta))
            else:
                subprocess.Popen(['xdg-open', str(ruta)])
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo abrir el Excel:\n{e}')

    def copiar_lista_compra(self):
        texto = self.txt_compra.get('1.0', 'end-1c').strip()
        if not texto or texto.startswith('Aún no hay ingredientes'):
            messagebox.showinfo('Aviso', 'La lista de la compra está vacía.')
            return
        self.root.clipboard_clear()
        self.root.clipboard_append(texto)
        self.root.update()
        messagebox.showinfo('Copiado', 'La lista de la compra se ha copiado al portapapeles.')

    def recargar_excel(self, mostrar_ok=True):
        try:
            self.platos = self.cargar_platos_excel()
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo leer la base de platos:\n{e}')
            return
        self.categorias = self.obtener_categorias()
        self.categorias_expandida = set(self.categorias)
        self.lbl_excel.config(text=f'Base actual: {self.excel_path}')
        self.refrescar_arbol_platos()
        self.refrescar_resumen()
        self.quitar_plato_seleccionado()
        if mostrar_ok:
            messagebox.showinfo('Base recargada', f'Se ha releído:\n{self.excel_path}')

    def generar_ingredientes(self):
        ingredientes_totales = []
        for dia in DIAS:
            for tipo in TIPOS:
                for slot in SLOTS:
                    plato = self.plan[dia][tipo][slot]
                    info = self.buscar_plato(plato)
                    if info and info['ingredientes']:
                        ingredientes_totales.extend(info['ingredientes'])
        conteo = {}
        for ing in ingredientes_totales:
            canonico = normalizar_texto_compra(ing)
            if canonico:
                conteo[canonico] = conteo.get(canonico, 0) + 1
        return conteo

    def refrescar_resumen(self):
        self.txt_resumen.delete('1.0', 'end')
        self.txt_compra.delete('1.0', 'end')
        for dia in DIAS:
            self.txt_resumen.insert('end', f'{dia.capitalize()}\n')
            for tipo in TIPOS:
                primero = self.plan[dia][tipo]['primero'] or '-'
                segundo = self.plan[dia][tipo]['segundo'] or '-'
                self.txt_resumen.insert('end', f'  - {tipo} · 1º: {primero}\n')
                self.txt_resumen.insert('end', f'  - {tipo} · 2º: {segundo}\n')
            self.txt_resumen.insert('end', '\n')
        conteo = self.generar_ingredientes()
        if conteo:
            for ing in sorted(conteo):
                n = conteo[ing]
                self.txt_compra.insert('end', f'- {ing}')
                if n > 1:
                    self.txt_compra.insert('end', f' (x{n})')
                self.txt_compra.insert('end', '\n')
        else:
            self.txt_compra.insert('end', 'Aún no hay ingredientes cargados o no se han seleccionado platos.\n')

    def construir_html(self):
        rows = []
        for tipo in TIPOS:
            for slot in SLOTS:
                etiqueta = f'{tipo.capitalize()} · {"1º" if slot == "primero" else "2º"}'
                celdas = ''
                for dia in DIAS:
                    celdas += f'<td>{html.escape(self.plan[dia][tipo][slot] or "")}</td>'
                rows.append(f'<tr><th>{etiqueta}</th>{celdas}</tr>')
        return f'''<!doctype html>
<html lang="es"><head><meta charset="utf-8"><title>Plan semanal comidas</title>
<style>
body {{ font-family: Arial, sans-serif; margin: 20px; color:#111; }}
h1 {{ margin-bottom: 8px; }}
table {{ width:100%; border-collapse: collapse; margin-top: 12px; }}
th, td {{ border:1px solid #ccc; padding:10px; text-align:center; vertical-align:top; }}
th {{ background:#f3f6fb; }}
.small {{ color:#555; font-size:13px; }}
@page {{ size: A4 landscape; margin: 12mm; }}
</style></head><body>
<h1>Plan semanal de comidas</h1>
<div class="small">Lunes a domingo · comidas y cenas · primer y segundo plato</div>
<table>
<thead><tr><th></th>{''.join(f'<th>{d.capitalize()}</th>' for d in DIAS)}</tr></thead>
<tbody>{''.join(rows)}</tbody>
</table>
</body></html>'''

    def exportar_html(self):
        fecha_actual = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        ruta_html = BASE / f'plan_semanal_{fecha_actual}.html'
        ruta_html.write_text(self.construir_html(), encoding='utf-8')
        messagebox.showinfo('HTML exportado', f'Se ha generado:\n{ruta_html}')

    def exportar_pdf(self):
        fecha_actual = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        pdf_path = BASE / f'plan_semanal_{fecha_actual}.pdf'
        html_path = pdf_path.with_suffix('.html')
        
        # Guardar temporalmente como HTML para que Chromium pueda leerlo y generar el PDF
        html_path.write_text(self.construir_html(), encoding='utf-8')
        try:
            subprocess.run([
                'chromium-browser', '--headless', '--disable-gpu', '--no-sandbox',
                f'--print-to-pdf={pdf_path}', f'file://{html_path}'
            ], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo exportar el PDF:\n{e}')
            return
        finally:
            if html_path.exists():
                html_path.unlink()
                
        messagebox.showinfo('PDF exportado', f'Se ha generado:\n{pdf_path}')


if __name__ == '__main__':
    root = tk.Tk()
    app = PlanificadorComidas(root)
    root.mainloop()
